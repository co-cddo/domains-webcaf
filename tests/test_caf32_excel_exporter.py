import os
import unittest

from openpyxl.worksheet.worksheet import Worksheet

from webcaf.webcaf.caf.routers import CAF32ExcelExporter


class CAF32ExcelExporterWithFixture(CAF32ExcelExporter):
    def get_framework_path(self) -> str:
        return os.path.join(os.path.dirname(__file__), "../frameworks", "cyber-assessment-framework-v3.2.yaml")


class TestCAF32ExcelExporter(unittest.TestCase):
    def setUp(self):
        # self.exporter = CAF32ExcelExporterWithFixture()
        self.exporter = CAF32ExcelExporter()
        self.wb = self.exporter.execute()
        # Uncomment to save the workbook for manual inspection
        current_dir = os.path.dirname(os.path.abspath(__file__))
        new_file_path = os.path.join(current_dir, "test.xlsx")
        self.wb.save(new_file_path)

    def test_pass(self):
        pass

    def test_elements_loaded_and_workbook_created(self):
        self.assertIsInstance(self.exporter.elements, list)
        self.assertGreater(len(self.exporter.elements), 0)
        self.assertIsNotNone(self.wb)

    def test_sheet_count_and_titles(self):
        # Expect one sheet per objective in the dummy fixture (A and B)
        sheet_titles = [ws.title for ws in self.wb.worksheets]
        # Titles use the format used by exporter: "CAF - Objective {code}" or similar
        self.assertTrue(any(title.endswith("Objective A") or title.endswith("Objective - A") for title in sheet_titles))
        self.assertTrue(any(title.endswith("Objective B") or title.endswith("Objective - B") for title in sheet_titles))

    def test_top_header_contents_on_first_sheet(self):
        ws: Worksheet = self.wb.worksheets[0]
        # Title merged across C..I at row 1 with the expected text
        self.assertEqual(ws.cell(row=1, column=1).value, "OFFICIAL SENSITIVE WHEN COMPLETED")
        # Resource Links header exists
        resource_links_row = None
        for r in range(1, 15):
            if ws.cell(row=r, column=1).value == "Supporting Resources":
                resource_links_row = r
                break
        self.assertIsNotNone(resource_links_row)
        # Specific link texts exist in subsequent rows
        texts = [ws.cell(row=resource_links_row + i, column=1).value for i in range(1, 4)]

        self.assertIn("Stage 3 self-assessment guidance:", texts)
        self.assertIn("NCSC CAF version 3.2:", texts)
        self.assertIn("WebCAF:", texts)

    def test_outcome_sections_and_validations_present(self):
        ws: Worksheet = self.wb.worksheets[1]
        # Find the first outcome header row by searching for a known pattern "A1.a -" or similar
        found_outcome_row = self._find_first_outcome_row(ws)
        self.assertIsNotNone(found_outcome_row, "Outcome header not found")
        # Column headers should be 3 rows below outcome header (one row for header, one for description)
        header_row = found_outcome_row + 3
        expected_headers = [
            "Justification (if applicable)",
            "Partially Achieved",
            "Answer",
            "Justification (if applicable)",
            "Not Achieved",
            "Answer",
            "Justification (if applicable)",
        ]
        actual_headers = [ws.cell(row=header_row, column=c).value for c in range(3, 10)]
        self.assertEqual(actual_headers, expected_headers)
        # Below headers there should be at least one indicator row with validation cells at D, F, H
        indicator_row = header_row + 1
        # The validation objects are registered on the worksheet; we can assert dataValidations exists
        self.assertTrue(hasattr(ws, "data_validations"))
        # Ensure the answer cells exist (they might be empty strings but should be addressable)
        for col in (4, 6, 8):
            self.assertIsNotNone(ws.cell(row=indicator_row, column=col))

    def test_indicator_fill_colours(self):
        # Validate fill colors for the first indicator line across achieved/partially/not columns
        ws: Worksheet = self.wb.worksheets[1]
        # Locate the first header row as in previous test
        found_outcome_row = self._find_first_outcome_row(ws, "A2.a")
        self.assertIsNotNone(
            found_outcome_row,
        )
        header_row = found_outcome_row + 3
        # Find the first indicator row that actually has any indicator text in C/E/G
        indicator_row = None
        scan_row = header_row + 1
        for r in range(scan_row, scan_row + 100):
            c_vals = [ws.cell(row=r, column=c).value for c in (1, 4, 7)]
            if any(isinstance(v, str) and v.strip() for v in c_vals):
                indicator_row = r
                break
        self.assertIsNotNone(indicator_row, "Could not locate a non-empty indicator row after headers")

        # Columns C, E, G should have colored fills for achieved/partially/not respectively
        fills = {
            1: "C6E2B3",  # green
            4: "FFFACD",  # yellow
            7: "FFB6C1",  # pink
        }
        for col, expected in fills.items():
            cell = ws.cell(row=indicator_row, column=col)
            fill = cell.fill
            self.assertIsNotNone(fill)
            # openpyxl may store color as ARGB or RGB or by index; normalize and compare suffix
            color_val = None
            if getattr(fill, "start_color", None) is not None:
                color_val = fill.start_color.rgb or fill.start_color.index
            if not color_val and getattr(fill, "fgColor", None) is not None:
                color_val = fill.fgColor.rgb or fill.fgColor.index
            self.assertIsNotNone(color_val, f"Could not read cell fill color at row {indicator_row}, col {col}")
            self.assertTrue(
                str(color_val).upper().endswith(expected), f"Expected fill {expected} at column {col}, got {color_val}"
            )

    def _find_first_outcome_row(self, ws: Worksheet, text_to_match: str = "A1.a") -> int | None:
        found_outcome_row = None
        for r in range(1, 200):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and (" - " in v) and (text_to_match in v):
                found_outcome_row = r
                break
        return found_outcome_row

    def _find_status_cells(self, worksheet):
        """Helper method to find all Contributing Outcome status cells in a worksheet"""
        status_cells = []
        for r in range(1, 500):  # Scan through enough rows to find all outcomes
            cell = worksheet.cell(row=r, column=1)
            if cell.value == "Contributing Outcome status:":
                # The status dropdown should be in column 2 on the same row
                status_cell = worksheet.cell(row=r, column=2)
                status_cells.append((r, status_cell))
        return status_cells

    def _check_cell_validation(self, worksheet, status_cell, row_num):
        """Helper method to check validation for a cell"""
        has_validation = False
        expected_options = None

        for dv in worksheet.data_validations.dataValidation:
            if status_cell.coordinate in dv.cells:
                has_validation = True
                # Check validation type
                self.assertEqual(
                    dv.type, "list", f"Cell {status_cell.coordinate} has wrong validation type in {worksheet.title}"
                )

                # Extract options from formula
                formula = dv.formula1.strip('"')
                options = formula.split(",")

                if len(options) == 2:
                    self.assertEqual(
                        options,
                        ["Achieved", "Not achieved"],
                        f"Incorrect options in {worksheet.title} at row {row_num}",
                    )
                    expected_options = ["Achieved", "Not achieved"]
                elif len(options) == 3:
                    self.assertEqual(
                        options,
                        ["Achieved", "Partially achieved", "Not achieved"],
                        f"Incorrect options in {worksheet.title} at row {row_num}",
                    )
                    expected_options = ["Achieved", "Partially achieved", "Not achieved"]
                else:
                    self.fail(f"Unexpected number of options: {options} in {worksheet.title} at row {row_num}")
                break

        # Ensure validation was found
        self.assertTrue(has_validation, f"No validation found for status cell at row {row_num} in {worksheet.title}")
        return expected_options

    def test_contributing_outcome_status_dropdowns(self):
        """Test that all Contributing Outcome status cells have the correct dropdown validation."""
        # Skip the guidance sheet (index 0)
        for sheet_idx in range(1, len(self.wb.worksheets)):
            ws = self.wb.worksheets[sheet_idx]
            self.assertTrue(ws.data_validations, f"No data validations found in sheet {ws.title}")

            # Find all contributing outcome status cells
            status_cells = self._find_status_cells(ws)
            self.assertTrue(len(status_cells) > 0, f"No 'Contributing Outcome status' cells found in sheet {ws.title}")

            # For each status cell, verify it has data validation and correct options
            for row_num, status_cell in status_cells:
                self._check_cell_validation(ws, status_cell, row_num)

    def _find_indicator_cells(self, worksheet):
        """Helper method to find indicator cells with Answer columns"""
        indicator_cells = []
        for r in range(1, 500):  # Scan through enough rows to find all indicators
            for col in [4, 6, 8]:  # Answer columns are D, F, H (4, 6, 8)
                cell = worksheet.cell(row=r, column=col)
                # Check if this cell has any data validation
                for dv in worksheet.data_validations.dataValidation:
                    if cell.coordinate in dv.cells:
                        indicator_cells.append((r, col, cell))
                        break
        return indicator_cells

    def test_indicator_answer_validations(self):
        """Test that all indicator answer cells have the correct True/False validation."""
        # Skip the guidance sheet (index 0)
        for sheet_idx in range(1, len(self.wb.worksheets)):
            ws = self.wb.worksheets[sheet_idx]
            self.assertTrue(ws.data_validations, f"No data validations found in sheet {ws.title}")

            # Find all indicator answer cells
            indicator_cells = self._find_indicator_cells(ws)
            self.assertTrue(len(indicator_cells) > 0, f"No indicator answer cells found in sheet {ws.title}")

            # For each indicator cell, verify it has data validation with True/False options
            for row_num, col_num, cell in indicator_cells:
                has_validation = False
                for dv in ws.data_validations.dataValidation:
                    if cell.coordinate in dv.cells:
                        has_validation = True
                        # Check validation type
                        self.assertEqual(
                            dv.type, "list", f"Cell {cell.coordinate} has wrong validation type in {ws.title}"
                        )

                        # Extract options from formula
                        formula = dv.formula1.strip('"')
                        options = formula.split(",")

                        # Indicator cells should have True/False options
                        self.assertEqual(
                            options,
                            ["True", "False"],
                            f"Incorrect options in {ws.title} at cell {cell.coordinate} (row {row_num}, col {col_num}). Expected True/False but got {options}",
                        )
                        break

                # Ensure validation was found
                self.assertTrue(
                    has_validation,
                    f"No validation found for indicator cell at row {row_num}, column {col_num} in {ws.title}",
                )


if __name__ == "__main__":
    unittest.main()
