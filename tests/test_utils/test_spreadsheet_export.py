import json
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from tests.test_views.base_view_test import BaseViewTest
from webcaf.webcaf.models import Assessment, Review
from webcaf.webcaf.utils.to_spreadsheet import export_assessment_to_excel

FIXTURE_PATH = Path(__file__).parent.parent / "fixtures" / "completed_assessment_base.json"
REVIEW_FIXTURE_PATH = Path(__file__).parent.parent / "fixtures" / "completed_review.json"


class TestCSVExport(BaseViewTest):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        all_systems = []

        for org in cls.org_map.values():
            for system in org["systems"].values():
                all_systems.append(system)

        with open(FIXTURE_PATH, "r") as f:
            base_assessment = json.load(f)

        with open(REVIEW_FIXTURE_PATH, "r") as f:
            base_review = json.load(f)

        # Make one assessment incomplete
        assessments_data = [base_assessment.copy() for _ in range(len(all_systems))]
        # We might want to edit reviews so create in the same way
        reviews_data = [base_review.copy() for _ in range(len(all_systems))]
        del assessments_data[-1]["D2.b"]

        cls.assessments = []
        cls.reviews = []
        for org in cls.org_map.values():
            for system in org["systems"].values():
                assessment = Assessment.objects.create(
                    system=system,
                    status="submitted",
                    assessment_period="25/26",
                    review_type="peer_review",
                    framework="caf32",
                    caf_profile="baseline",
                    assessments_data=assessments_data.pop(0),
                )
                cls.assessments.append(assessment)

                review = Review.objects.create(
                    assessment=assessment,
                    status="completed",
                    review_data=reviews_data.pop(0),
                )
                cls.reviews.append(review)

    def test_we_have_some_assessments(self):
        assert len(self.assessments) > 0

    def test_igp_numbering_restarts_per_category(self):
        assessment = self.assessments[0]

        excel_bytes = export_assessment_to_excel(assessment)
        assert excel_bytes is not None

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Indicator Level"]

        igp_labels = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            contributing_outcome = row[0]
            igp_label = row[1]
            if contributing_outcome and igp_label:
                igp_labels.append(igp_label)

        a1a_not_achieved = [label for label in igp_labels if label.startswith("A1.a Not achieved statement")]
        a1a_achieved = [label for label in igp_labels if label.startswith("A1.a Achieved statement")]

        assert "A1.a Not achieved statement 1" in a1a_not_achieved
        assert "A1.a Not achieved statement 2" in a1a_not_achieved
        assert "A1.a Not achieved statement 3" in a1a_not_achieved
        assert "A1.a Not achieved statement 4" in a1a_not_achieved

        assert "A1.a Achieved statement 1" in a1a_achieved
        assert "A1.a Achieved statement 2" in a1a_achieved
        assert "A1.a Achieved statement 3" in a1a_achieved
        assert "A1.a Achieved statement 4" in a1a_achieved

        a1c_not_achieved = [label for label in igp_labels if label.startswith("A1.c Not achieved statement")]
        a1c_achieved = [label for label in igp_labels if label.startswith("A1.c Achieved statement")]

        assert "A1.c Not achieved statement 1" in a1c_not_achieved
        assert "A1.c Not achieved statement 2" in a1c_not_achieved
        assert "A1.c Not achieved statement 5" in a1c_not_achieved

        assert "A1.c Achieved statement 1" in a1c_achieved
        assert "A1.c Achieved statement 2" in a1c_achieved
        assert "A1.c Achieved statement 4" in a1c_achieved

    def test_indicator_values_displayed_as_y_or_n(self):
        assessment = self.assessments[0]

        excel_bytes = export_assessment_to_excel(assessment)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Indicator Level"]

        for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
            self_assessment = row[3]
            review = row[5]
            assert self_assessment in ["Y", "N"]
            assert review in ["Y", "N"]

    # Exports a timestamped .xlsx file to the project root
    def test_export_single_assessment_to_file(self):
        assessment = self.assessments[0]

        excel_bytes = export_assessment_to_excel(assessment)
        assert excel_bytes is not None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path(__file__).parent.parent.parent / f"assessment_export_{timestamp}.xlsx"

        with open(output_path, "wb") as f:
            f.write(excel_bytes)

        assert output_path.exists()
        assert output_path.stat().st_size > 0
