"""
Tests for webcaf/webcaf/utils/to_spreadsheet.py

This module contains tests for the spreadsheet export functionality,
specifically testing the peer_review specific logic changes:
- Review type title casing in metadata tab
- Conditional "Review comments" column based on review type
- Conditional sheet name for recommendations tab
- Conditional "Risk" and "Recommendation number" columns based on review type
"""
import json
from io import BytesIO
from pathlib import Path

from django.test import TestCase
from openpyxl import load_workbook

from webcaf.webcaf.models import Assessment, Review
from webcaf.webcaf.utils.to_spreadsheet import review_to_excel

FIXTURE_PATH = Path(__file__).parent.parent / "fixtures" / "completed_assessment_base.json"
REVIEW_FIXTURE_PATH = Path(__file__).parent.parent / "fixtures" / "completed_review.json"


class TestReviewToExcelPeerReview(TestCase):
    """Tests for review_to_excel with peer_review review type."""

    @classmethod
    def setUpTestData(cls):
        """Set up test data with both independent and peer review assessments."""
        from django.contrib.auth.models import User

        from webcaf.webcaf.models import Organisation, System, UserProfile

        cls.org = Organisation.objects.create(name="Test Organisation")
        cls.system = System.objects.create(name="Test System", organisation=cls.org)
        cls.system_2 = System.objects.create(name="Test System 2", organisation=cls.org)
        cls.system_3 = System.objects.create(name="Test System 3", organisation=cls.org)

        cls.user = User.objects.create_user(username="test@test.gov.uk", email="test@test.gov.uk")
        UserProfile.objects.create(user=cls.user, organisation=cls.org, role="cyber_advisor")

        with open(FIXTURE_PATH, "r") as f:
            base_assessment = json.load(f)

        with open(REVIEW_FIXTURE_PATH, "r") as f:
            base_review = json.load(f)

        # Create independent review assessment
        cls.independent_assessment = Assessment.objects.create(
            system=cls.system,
            status="submitted",
            assessment_period="25/26",
            review_type="independent",
            framework="caf32",
            caf_profile="baseline",
            assessments_data=base_assessment,
        )

        cls.other_assessment = Assessment.objects.create(
            system=cls.system_3,
            status="submitted",
            assessment_period="25/26",
            review_type="independent",
            framework="caf32",
            caf_profile="baseline",
            assessments_data=base_assessment,
        )

        cls.independent_review = Review.objects.create(
            assessment=cls.independent_assessment,
            status="completed",
            review_data=base_review,
        )

        # Create peer review assessment
        cls.peer_review_assessment = Assessment.objects.create(
            system=cls.system_2,
            status="submitted",
            assessment_period="25/26",
            review_type="peer_review",
            framework="caf32",
            caf_profile="baseline",
            assessments_data=base_assessment,
        )

        cls.peer_review = Review.objects.create(
            assessment=cls.peer_review_assessment,
            status="completed",
            review_data=base_review,
        )

    def test_review_to_excel_returns_bytes(self):
        """Test that review_to_excel returns bytes for completed review."""
        result = review_to_excel(self.independent_review)
        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 0)

    def test_review_to_excel_returns_none_for_incomplete_review(self):
        """Test that review_to_excel returns None for incomplete review."""
        incomplete_review = Review.objects.create(
            assessment=self.other_assessment,
            status="draft",
            review_data={},
        )
        result = review_to_excel(incomplete_review)
        self.assertIsNone(result)

    def test_metadata_tab_review_type_title_casing(self):
        """Test that review type is title cased in metadata tab."""
        # Test independent review
        excel_bytes = review_to_excel(self.independent_review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Review details"]

        # Find the review type row
        review_type_value = None
        for row in ws.iter_rows(values_only=True):
            if row[0] and "Review type" in str(row[0]):
                review_type_value = row[1]
                break

        # Should be title cased (e.g., "Independent Assurance Review" not "independent")
        self.assertIsNotNone(review_type_value)
        self.assertEqual(review_type_value, "Independent Assurance Review")

        # Test peer review
        excel_bytes = review_to_excel(self.peer_review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Review details"]

        review_type_value = None
        for row in ws.iter_rows(values_only=True):
            if row[0] and "Review type" in str(row[0]):
                review_type_value = row[1]
                break

        # Should be title cased (e.g., "Peer Review" not "peer_review")
        self.assertIsNotNone(review_type_value)
        self.assertEqual(review_type_value, "Peer Review")

    def test_igp_tab_has_review_comments_for_independent_review(self):
        """Test that IGP tab includes 'Review comments' column for independent review."""
        excel_bytes = review_to_excel(self.independent_review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["IGPs"]

        # Get header row
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Review comments", headers)

    def test_igp_tab_no_review_comments_for_peer_review(self):
        """Test that IGP tab does NOT include 'Review comments' column for peer review."""
        excel_bytes = review_to_excel(self.peer_review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["IGPs"]

        # Get header row
        headers = [cell.value for cell in ws[1]]
        self.assertNotIn("Review comments", headers)

    def test_igp_tab_column_count_differs_by_review_type(self):
        """Test that IGP tab has different column counts for different review types."""
        # Independent review should have 7 columns
        excel_bytes = review_to_excel(self.independent_review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["IGPs"]
        independent_cols = len([cell.value for cell in ws[1] if cell.value])

        # Peer review should have 6 columns (no review comments)
        excel_bytes = review_to_excel(self.peer_review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["IGPs"]
        peer_cols = len([cell.value for cell in ws[1] if cell.value])

        # Independent should have one more column than peer review
        self.assertEqual(independent_cols, peer_cols + 1)

    def test_recommendations_tab_sheet_name_independent_review(self):
        """Test that recommendations tab is named 'Risks and recommendations' for independent review."""
        excel_bytes = review_to_excel(self.independent_review)
        wb = load_workbook(BytesIO(excel_bytes))

        sheet_names = wb.sheetnames
        self.assertIn("Risks and recommendations", sheet_names)

    def test_recommendations_tab_sheet_name_peer_review(self):
        """Test that recommendations tab is named 'Recommendations' for peer review."""
        excel_bytes = review_to_excel(self.peer_review)
        wb = load_workbook(BytesIO(excel_bytes))

        sheet_names = wb.sheetnames
        self.assertIn("Recommendations", sheet_names)
        # Peer review should NOT have the "Risks and recommendations" sheet
        self.assertNotIn("Risks and recommendations", sheet_names)

    def test_recommendations_tab_has_risk_columns_for_independent_review(self):
        """Test that recommendations tab includes 'Risk' and 'Recommendation number' for independent review."""
        excel_bytes = review_to_excel(self.independent_review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Risks and recommendations"]

        # Get header row
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Risk", headers)
        self.assertIn("Recommendation number", headers)

    def test_recommendations_tab_no_risk_columns_for_peer_review(self):
        """Test that recommendations tab does NOT include 'Risk' and 'Recommendation number' for peer review."""
        excel_bytes = review_to_excel(self.peer_review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Recommendations"]

        # Get header row
        headers = [cell.value for cell in ws[1]]
        self.assertNotIn("Risk", headers)
        self.assertNotIn("Recommendation number", headers)

    def test_recommendations_tab_column_count_differs_by_review_type(self):
        """Test that recommendations tab has different column counts for different review types."""
        # Independent review should have 6 columns
        excel_bytes = review_to_excel(self.independent_review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Risks and recommendations"]
        independent_cols = len([cell.value for cell in ws[1] if cell.value])

        # Peer review should have 4 columns (no Risk and Recommendation number)
        excel_bytes = review_to_excel(self.peer_review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Recommendations"]
        peer_cols = len([cell.value for cell in ws[1] if cell.value])

        # Independent should have two more columns than peer review
        self.assertEqual(independent_cols, peer_cols + 2)

    def test_all_sheets_present_for_both_review_types(self):
        """Test that all required sheets are present for both review types."""
        required_sheets = ["Review details", "IGPs", "Contributing outcomes"]

        # Test independent review
        excel_bytes = review_to_excel(self.independent_review)
        wb = load_workbook(BytesIO(excel_bytes))
        for sheet in required_sheets:
            self.assertIn(sheet, wb.sheetnames, f"Sheet '{sheet}' missing for independent review")

        # Test peer review
        excel_bytes = review_to_excel(self.peer_review)
        wb = load_workbook(BytesIO(excel_bytes))
        for sheet in required_sheets:
            self.assertIn(sheet, wb.sheetnames, f"Sheet '{sheet}' missing for peer review")

    def test_recommendations_tab_common_columns_present(self):
        """Test that common columns are present in recommendations tab for both review types."""
        common_columns = ["Contributing outcome", "Target CAF profile", "Risk number", "Recommendation"]

        # Test independent review
        excel_bytes = review_to_excel(self.independent_review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Risks and recommendations"]
        headers = [cell.value for cell in ws[1]]
        for col in common_columns:
            self.assertIn(col, headers)

        # Test peer review
        excel_bytes = review_to_excel(self.peer_review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Recommendations"]
        headers = [cell.value for cell in ws[1]]
        for col in common_columns:
            self.assertIn(col, headers)


class TestReviewToExcelOtherReviewTypes(TestCase):
    """Tests for review_to_excel with other review types."""

    @classmethod
    def setUpTestData(cls):
        """Set up test data with different review types."""
        from django.contrib.auth.models import User

        from webcaf.webcaf.models import Organisation, System, UserProfile

        cls.org = Organisation.objects.create(name="Test Organisation")
        cls.system = System.objects.create(name="Test System", organisation=cls.org)
        cls.system_2 = System.objects.create(name="Test System 2", organisation=cls.org)

        cls.user = User.objects.create_user(username="test@test.gov.uk", email="test@test.gov.uk")
        UserProfile.objects.create(user=cls.user, organisation=cls.org, role="cyber_advisor")

        with open(FIXTURE_PATH, "r") as f:
            base_assessment = json.load(f)

        with open(REVIEW_FIXTURE_PATH, "r") as f:
            base_review = json.load(f)

        # Create assessments with different review types
        cls.review_types = ["independent", "peer_review"]
        cls.reviews = {}

        for review_type, system in zip(cls.review_types, [cls.system, cls.system_2]):
            assessment = Assessment.objects.create(
                system=system,
                status="submitted",
                assessment_period="25/26",
                review_type=review_type,
                framework="caf32",
                caf_profile="baseline",
                assessments_data=base_assessment,
            )

            review = Review.objects.create(
                assessment=assessment,
                status="completed",
                review_data=base_review,
            )
            cls.reviews[review_type] = review

    def test_non_peer_review_types_have_review_comments_column(self):
        """Test that non-peer review types have 'Review comments' column."""
        for review_type in ["independent"]:
            review = self.reviews[review_type]
            excel_bytes = review_to_excel(review)
            wb = load_workbook(BytesIO(excel_bytes))
            ws = wb["IGPs"]
            headers = [cell.value for cell in ws[1]]
            self.assertIn("Review comments", headers, f"Missing for {review_type}")

    def test_non_peer_review_types_have_risk_columns(self):
        """Test that non-peer review types have 'Risk' and 'Recommendation number' columns."""
        for review_type in ["independent"]:
            review = self.reviews[review_type]
            excel_bytes = review_to_excel(review)
            wb = load_workbook(BytesIO(excel_bytes))
            ws = wb["Risks and recommendations"]
            headers = [cell.value for cell in ws[1]]
            self.assertIn("Risk", headers, f"Missing for {review_type}")
            self.assertIn("Recommendation number", headers, f"Missing for {review_type}")

    def test_peer_review_does_not_have_review_comments_column(self):
        """Test that peer_review does NOT have 'Review comments' column."""
        review = self.reviews["peer_review"]
        excel_bytes = review_to_excel(review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["IGPs"]
        headers = [cell.value for cell in ws[1]]
        self.assertNotIn("Review comments", headers)

    def test_peer_review_does_not_have_risk_columns(self):
        """Test that peer_review does NOT have 'Risk' and 'Recommendation number' columns."""
        review = self.reviews["peer_review"]
        excel_bytes = review_to_excel(review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Recommendations"]
        headers = [cell.value for cell in ws[1]]
        self.assertNotIn("Risk", headers)
        self.assertNotIn("Recommendation number", headers)


class TestReviewToExcelDataIntegrity(TestCase):
    """Tests to ensure data integrity is maintained with the peer_review changes."""

    @classmethod
    def setUpTestData(cls):
        """Set up test data."""
        from django.contrib.auth.models import User

        from webcaf.webcaf.models import Organisation, System, UserProfile

        cls.org = Organisation.objects.create(name="Test Organisation")
        cls.system = System.objects.create(name="Test System", organisation=cls.org)
        cls.user = User.objects.create_user(username="test@test.gov.uk", email="test@test.gov.uk")
        UserProfile.objects.create(user=cls.user, organisation=cls.org, role="cyber_advisor")

        with open(FIXTURE_PATH, "r") as f:
            base_assessment = json.load(f)

        with open(REVIEW_FIXTURE_PATH, "r") as f:
            base_review = json.load(f)

        cls.assessment = Assessment.objects.create(
            system=cls.system,
            status="submitted",
            assessment_period="25/26",
            review_type="peer_review",
            framework="caf32",
            caf_profile="baseline",
            assessments_data=base_assessment,
        )

        cls.review = Review.objects.create(
            assessment=cls.assessment,
            status="completed",
            review_data=base_review,
        )

    def test_igp_data_rows_have_correct_column_count(self):
        """Test that data rows in IGP tab have the correct column count for peer review."""
        excel_bytes = review_to_excel(self.review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["IGPs"]

        # Header should have 6 columns for peer review
        header_count = len([cell.value for cell in ws[1] if cell.value])
        self.assertEqual(header_count, 6)

        # Check a few data rows have the same column count
        for row_idx in range(2, min(10, ws.max_row + 1)):
            row = [cell.value for cell in ws[row_idx]]
            # Non-empty cells should match header count
            non_empty_count = len([cell for cell in row if cell is not None])
            # Allow for some cells to be empty, but structure should be consistent
            self.assertLessEqual(non_empty_count, header_count)

    def test_recommendations_data_rows_have_correct_column_count(self):
        """Test that data rows in recommendations tab have the correct column count for peer review."""
        excel_bytes = review_to_excel(self.review)
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Recommendations"]

        # Header should have 4 columns for peer review
        header_count = len([cell.value for cell in ws[1] if cell.value])
        self.assertEqual(header_count, 4)

        # Check data rows if they exist
        if ws.max_row > 1:
            for row_idx in range(2, min(10, ws.max_row + 1)):
                row = [cell.value for cell in ws[row_idx]]
                non_empty_count = len([cell for cell in row if cell is not None])
                self.assertLessEqual(non_empty_count, header_count)

    def test_excel_file_is_valid(self):
        """Test that the generated Excel file is valid and can be loaded."""
        excel_bytes = review_to_excel(self.review)

        # Should be able to load without errors
        wb = load_workbook(BytesIO(excel_bytes))

        # Should have expected sheets
        self.assertIn("Review details", wb.sheetnames)
        self.assertIn("IGPs", wb.sheetnames)
        self.assertIn("Contributing outcomes", wb.sheetnames)
        self.assertIn("Recommendations", wb.sheetnames)

        # Should be able to iterate through all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Should be able to access max_row
            _ = ws.max_row
            # Should be able to iterate through rows
            for _ in ws.iter_rows(values_only=True):
                pass
