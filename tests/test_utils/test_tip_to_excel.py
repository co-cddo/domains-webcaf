import unittest
from io import BytesIO
from unittest.mock import MagicMock

from openpyxl import load_workbook

from webcaf.webcaf.utils.review import Recommendation, RecommendationGroup
from webcaf.webcaf.utils.to_spreadsheet import tip_to_excel


class TestTipToExcel(unittest.TestCase):
    def setUp(self):
        self.mock_tip = MagicMock()
        self.mock_review = MagicMock()
        self.mock_assessment = MagicMock()

        self.mock_tip.review = self.mock_review
        self.mock_review.assessment = self.mock_assessment

        self.mock_assessment.system.organisation.name = "Big organisation"
        self.mock_assessment.system.name = "Medium system"
        self.mock_assessment.assessment_period = "25/26"
        self.mock_assessment.review_type = "independent"
        self.mock_assessment.framework = "caf32"
        self.mock_assessment.caf_profile = "baseline"

        self.mock_tip.is_submitted = True
        self.mock_tip.is_approved = False

    def test_tip_to_excel_independent_planned_action(self):
        # Setup context
        rec1 = Recommendation(
            id="REC1",
            title="Risk 1",
            text="Recommendation text 1",
            objective="OBJ1",
            outcome="OUT1",
            outcome_title="Outcome Title 1",
        )
        group1 = RecommendationGroup(title="Risk 1", recommendations=[rec1], group_index=1)

        action1 = MagicMock()
        action1.action_type = "action_planned"
        action1.recommendation_category = "priority"
        action1.recommendation_reviewed = "yes"
        action1.action_details = {
            "action_taken_description": "Will do X",
            "action_owner": "Owner 1",
            "resources_available": "Yes",
            "budget_available": "Yes",
            "target_date_provided": "yes",
            "target_day_day": "01",
            "target_day_month": "01",
            "target_day_year": "2026",
        }

        context = {
            "priority_recommendations": [(rec1, group1, action1)],
            "other_recommendations": [],
        }

        excel_bytes = tip_to_excel(self.mock_tip, context)
        self.assertIsNotNone(excel_bytes)

        wb = load_workbook(BytesIO(excel_bytes))

        # Check Review details sheet
        ws_details = wb["Review details"]
        self.assertEqual(ws_details["B1"].value, "Big organisation")
        self.assertEqual(ws_details["B2"].value, "Medium system")
        self.assertEqual(ws_details["B7"].value, "Submitted")

        # Check Recommendations and actions sheet
        ws_actions = wb["Recommendations and actions"]

        # Row 2: action1
        row2 = [cell.value for cell in ws_actions[2]]
        self.assertEqual(row2[0], "Priority")
        self.assertEqual(row2[1], "OUT1 Outcome Title 1")
        self.assertEqual(row2[2], "RP1 — Risk 1")
        self.assertEqual(row2[3], "REC1 - Recommendation text 1")
        self.assertEqual(row2[4], "Yes")
        self.assertEqual(row2[5], "Action planned")
        self.assertEqual(row2[6], "N/A")
        self.assertEqual(row2[7], "Will do X")
        self.assertEqual(row2[8], "Owner 1")
        self.assertEqual(row2[9], "Yes")
        self.assertEqual(row2[10], "Yes")
        self.assertEqual(row2[11], "01/01/2026")
        self.assertEqual(row2[12], "N/A")

    def test_tip_to_excel_independent_not_planned_action(self):
        rec2 = Recommendation(
            id="REC2",
            title="Risk 2",
            text="Recommendation text 2",
            objective="OBJ1",
            outcome="OUT1",
            outcome_title="Outcome Title 1",
        )
        group2 = RecommendationGroup(title="Risk 2", recommendations=[rec2], group_index=2)

        action2 = MagicMock()
        action2.action_type = "action_not_planned"
        action2.recommendation_category = "other"
        action2.recommendation_reviewed = "yes"
        action2.action_details = {
            "action_not_planned_reason": "No need",
        }

        context = {
            "priority_recommendations": [],
            "other_recommendations": [(rec2, group2, action2)],
        }

        excel_bytes = tip_to_excel(self.mock_tip, context)
        self.assertIsNotNone(excel_bytes)

        wb = load_workbook(BytesIO(excel_bytes))
        ws_actions = wb["Recommendations and actions"]

        # Row 2: action2
        row2 = [cell.value for cell in ws_actions[2]]
        self.assertEqual(row2[0], "Other")
        self.assertEqual(row2[1], "OUT1 Outcome Title 1")
        self.assertEqual(row2[2], "RO2 — Risk 2")
        self.assertEqual(row2[3], "REC2 - Recommendation text 2")
        self.assertEqual(row2[4], "Yes")
        self.assertEqual(row2[5], "No action planned")
        self.assertEqual(row2[6], "No need")
        self.assertEqual(row2[7], "N/A")
        self.assertEqual(row2[8], "N/A")
        self.assertEqual(row2[9], "N/A")
        self.assertEqual(row2[10], "N/A")
        self.assertEqual(row2[11], "N/A")
        self.assertEqual(row2[12], "N/A")

    def test_tip_to_excel_peer_review(self):
        self.mock_assessment.review_type = "peer_review"

        rec = Recommendation(
            id="REC3",
            title="Risk 3",
            text="Recommendation text 3",
            objective="OBJ1",
            outcome="OUT1",
            outcome_title="Outcome Title 1",
        )
        group = RecommendationGroup(title="Risk 3", recommendations=[rec], group_index=3)

        action = MagicMock()
        action.action_type = "action_planned"
        action.recommendation_category = "priority"
        action.recommendation_reviewed = "no"
        action.action_details = {
            "action_taken_description": "Will do Y",
            "action_owner": "Owner 2",
            "target_date_provided": "no",
            "target_date_unavailable_reason": "Pending budget",
        }

        context = {
            "priority_recommendations": [(rec, group, action)],
            "other_recommendations": [],
        }

        excel_bytes = tip_to_excel(self.mock_tip, context)
        self.assertIsNotNone(excel_bytes)

        wb = load_workbook(BytesIO(excel_bytes))
        ws_actions = wb["Recommendations and actions"]

        row2 = [cell.value for cell in ws_actions[2]]
        self.assertEqual(row2[0], "Priority")
        self.assertEqual(row2[1], "OUT1 Outcome Title 1")
        # Since Associated risk is missing, indices shift
        self.assertEqual(row2[2], "REC3 - Recommendation text 3")
        self.assertEqual(row2[3], "No")
        self.assertEqual(row2[4], "Action planned")
        self.assertEqual(row2[5], "N/A")
        self.assertEqual(row2[6], "Will do Y")
        self.assertEqual(row2[7], "Owner 2")
        self.assertEqual(row2[10], "No target date")
        self.assertEqual(row2[11], "Pending budget")

    def test_tip_to_excel_draft_status(self):
        self.mock_tip.is_submitted = False
        self.mock_tip.is_approved = False

        context = {
            "priority_recommendations": [],
            "other_recommendations": [],
        }

        excel_bytes = tip_to_excel(self.mock_tip, context)
        wb = load_workbook(BytesIO(excel_bytes))
        ws_details = wb["Review details"]
        self.assertEqual(ws_details["B7"].value, "Draft")
