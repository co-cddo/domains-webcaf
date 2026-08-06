import unittest
from io import BytesIO

from openpyxl import Workbook, load_workbook

from webcaf.webcaf.utils.excel_exporter import build_assessment_template_workbook
from webcaf.webcaf.utils.excel_importer import (
    JSON_MAP_SHEET_NAME,
    ExcelImportError,
    excel_to_assessment_json,
    excel_to_assessment_json_with_raw,
)

FRAMEWORK = {
    "objectives": {
        "A": {
            "code": "A",
            "title": "Managing security risk",
            "description": "Objective A description",
            "principles": {
                "A1": {
                    "code": "A1",
                    "title": "Governance",
                    "description": "Principle A1 description",
                    "outcomes": {
                        "A1.a": {
                            "code": "A1.a",
                            "title": "Board Direction",
                            "description": "Outcome A1.a description",
                            "indicators": {
                                "achieved": {
                                    "A1.a.1": {"description": "First achieved indicator"},
                                    "A1.a.2": {"description": "Second achieved indicator"},
                                },
                                "partially-achieved": {
                                    "A1.a.3": {"description": "Partially achieved indicator"},
                                },
                                "not-achieved": {
                                    "A1.a.4": {"description": "Not achieved indicator"},
                                },
                            },
                        },
                        "A1.b": {
                            "code": "A1.b",
                            "title": "Assurance",
                            "description": "Outcome A1.b description",
                            "indicators": {
                                "achieved": {"A1.b.1": {"description": "Only achieved indicator"}},
                                "not-achieved": {"A1.b.2": {"description": "Only not-achieved indicator"}},
                            },
                        },
                    },
                }
            },
        }
    }
}


class TestCAF32ExcelImporter(unittest.TestCase):
    def _workbook_file(self, values: dict) -> BytesIO:
        """Build the template for FRAMEWORK, fill the cells mapped to the given
        json paths with the given values, and return it as an uploadable file."""
        wb = build_assessment_template_workbook(FRAMEWORK)
        cell_map = self._cell_map(wb)
        for json_path, value in values.items():
            ws_title, coordinate = cell_map[json_path]
            wb[ws_title][coordinate] = value
        return self._save(wb)

    @staticmethod
    def _cell_map(wb: Workbook) -> dict:
        map_ws = wb[JSON_MAP_SHEET_NAME]
        return {
            json_path: (visible_sheet, visible_cell)
            for visible_sheet, visible_cell, json_path, _value_type, _required in map_ws.iter_rows(
                min_row=2, values_only=True
            )
        }

    @staticmethod
    def _save(wb: Workbook) -> BytesIO:
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    COMPLETE_ANSWERS = {
        "/A1.a/indicators/achieved_A1.a.1": "agreed",
        "/A1.a/indicators/achieved_A1.a.2": "not_true_no_justification",
        "/A1.a/indicators/partially-achieved_A1.a.3": "not_true_have_justification",
        "/A1.a/indicators/not-achieved_A1.a.4": "true_have_justification",
        "/A1.a/confirmation/outcome_status": "Partially achieved",
        "/A1.a/confirmation/confirm_outcome_confirm_comment": "Evidence for A1.a",
        "/A1.b/indicators/achieved_A1.b.1": "agreed",
        "/A1.b/indicators/not-achieved_A1.b.2": "not_true_no_justification",
        "/A1.b/confirmation/outcome_status": "Achieved",
    }

    def test_round_trip_produces_expected_json(self):
        data = excel_to_assessment_json(self._workbook_file(self.COMPLETE_ANSWERS))
        self.assertEqual(
            data,
            {
                "A1.a": {
                    "indicators": {
                        "achieved_A1.a.1": True,
                        "achieved_A1.a.2": False,
                        "partially-achieved_A1.a.3": False,
                        "not-achieved_A1.a.4": True,
                    },
                    "confirmation": {
                        "outcome_status": "Partially achieved",
                        "confirm_outcome_confirm_comment": "Evidence for A1.a",
                        "confirm_outcome": "confirm",
                    },
                },
                "A1.b": {
                    "indicators": {
                        "achieved_A1.b.1": True,
                        "not-achieved_A1.b.2": False,
                    },
                    "confirmation": {
                        "outcome_status": "Achieved",
                        # Blank comment cell is optional and imported as empty text
                        "confirm_outcome_confirm_comment": "",
                        "confirm_outcome": "confirm",
                    },
                },
            },
        )

    def test_blank_indicator_answers_import_as_false(self):
        answers = {key: value for key, value in self.COMPLETE_ANSWERS.items() if "indicators" not in key}
        data = excel_to_assessment_json(self._workbook_file(answers))
        self.assertEqual(
            data["A1.a"]["indicators"],
            {
                "achieved_A1.a.1": False,
                "achieved_A1.a.2": False,
                "partially-achieved_A1.a.3": False,
                "not-achieved_A1.a.4": False,
            },
        )

    def test_blank_required_status_raises_on_strict_import(self):
        answers = {key: value for key, value in self.COMPLETE_ANSWERS.items() if key != "/A1.b/confirmation/outcome_status"}
        with self.assertRaises(ExcelImportError) as ctx:
            excel_to_assessment_json(self._workbook_file(answers))
        self.assertIn("Required cell", str(ctx.exception))

    def test_blank_required_status_allowed_in_preview(self):
        answers = {key: value for key, value in self.COMPLETE_ANSWERS.items() if key != "/A1.b/confirmation/outcome_status"}
        data, raw_rows = excel_to_assessment_json_with_raw(self._workbook_file(answers))
        self.assertNotIn("outcome_status", data["A1.b"]["confirmation"])
        blank_status_rows = [
            row for row in raw_rows if row[0] == "/A1.b/confirmation/outcome_status"
        ]
        self.assertEqual(blank_status_rows, [("/A1.b/confirmation/outcome_status", "outcome_status", True, None)])

    def test_invalid_indicator_answer_raises(self):
        answers = {**self.COMPLETE_ANSWERS, "/A1.a/indicators/achieved_A1.a.1": "maybe"}
        with self.assertRaises(ExcelImportError) as ctx:
            excel_to_assessment_json(self._workbook_file(answers))
        self.assertIn("invalid indicator answer", str(ctx.exception))

    def test_invalid_outcome_status_raises(self):
        answers = {**self.COMPLETE_ANSWERS, "/A1.a/confirmation/outcome_status": "Sort of achieved"}
        with self.assertRaises(ExcelImportError) as ctx:
            excel_to_assessment_json(self._workbook_file(answers))
        self.assertIn("invalid outcome status", str(ctx.exception))

    def test_workbook_without_map_sheet_raises(self):
        with self.assertRaises(ExcelImportError) as ctx:
            excel_to_assessment_json(self._save(Workbook()))
        self.assertIn(f"missing the {JSON_MAP_SHEET_NAME} sheet", str(ctx.exception))

    def test_workbook_with_invalid_map_header_raises(self):
        wb = build_assessment_template_workbook(FRAMEWORK)
        wb[JSON_MAP_SHEET_NAME]["A1"] = "bogus"
        with self.assertRaises(ExcelImportError) as ctx:
            excel_to_assessment_json(self._save(wb))
        self.assertIn("invalid header row", str(ctx.exception))

    def test_map_sheet_is_hidden_in_exported_template(self):
        wb = load_workbook(self._workbook_file({}))
        self.assertEqual(wb[JSON_MAP_SHEET_NAME].sheet_state, "veryHidden")


if __name__ == "__main__":
    unittest.main()
