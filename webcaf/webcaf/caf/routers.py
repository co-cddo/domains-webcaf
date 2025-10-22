import logging
import os
from abc import abstractmethod
from typing import Any, Generator, Optional

import yaml
from django.conf import settings
from django.urls import path, reverse_lazy
from django.utils.text import slugify
from django.views.generic import FormView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from webcaf import urls
from webcaf.webcaf.abcs import FrameworkRouter
from webcaf.webcaf.caf.views.factory import create_form_view
from webcaf.webcaf.forms.factory import create_form

from .field_providers import (
    FieldProvider,
    OutcomeConfirmationFieldProvider,
    OutcomeIndicatorsFieldProvider,
)

FrameworkValue = str | dict | int | None

FormViewClass = type[FormView]

CAF32Element = dict[str, Any]


class CAFLoader(FrameworkRouter):
    """
    Represents a loader for a specific framework, responsible for loading, reading, and traversing
    a hierarchical framework structure defined in YAML format.

    The class is designed to provide an interface for managing a hierarchical framework structure,
    offering functionality to load the framework from a file, traverse its structure, and retrieve
    specific elements. Subclasses are required to provide specific implementations for framework
    path and ID retrieval. The traversal is operation-specific, allowing for flexibility in
    defining stages like "indicators" or "confirmation".

    :ivar framework: Dictionary containing the entire framework structure as parsed from
        the YAML file.
    :type framework: dict
    :ivar elements: List of all framework elements extracted and traversed from the
        framework structure.
    :type elements: list
    """

    def __init__(self) -> None:
        self.framework: CAF32Element = {}
        self.elements: list[CAF32Element] = []
        self._read()

    @abstractmethod
    def get_framework_path(self) -> str:
        """
        Needs to be implemented by subclasses
        :return:
        """

    @abstractmethod
    def get_framework_id(self) -> str:
        """
        Needs to be implemented by subclasses
        :return:
        """

    def _read(self) -> None:
        with open(self.get_framework_path(), "r") as file:
            self.framework = yaml.safe_load(file)
            self.elements = list(self._traverse_framework())

    def _traverse_framework(self) -> Generator[CAF32Element, None, None]:
        """
        Traverse the framework structure and yield those elements requiring their own
        page in a single sequence.
        """
        for objective_code, objective in self.framework.get("objectives", {}).items():
            objective_ = {
                # Add the dictionary taken from the YAML first so that our code value
                # is set from the dict key and not the value *within* the dict. We
                # can probably remove the code attributes from the YAML
                **objective,
                "type": "objective",
                "code": objective_code,
                "short_name": f"{self.get_framework_id()}_objective_{objective_code}",
                "parent": None,
            }
            yield objective_
            for principle_code, principle in objective.get("principles", {}).items():
                principle_ = {
                    **principle,
                    "type": "principle",
                    "code": principle_code,
                    "short_name": f"{self.get_framework_id()}_principle_{principle_code}",
                    "parent": objective_,
                }
                yield principle_
                for outcome_code, outcome in principle.get("outcomes", {}).items():
                    outcome_ = {
                        **outcome,
                        "type": "outcome",
                        "code": outcome_code,
                        "short_name": f"{self.get_framework_id()}_indicators_{outcome_code}",
                        "parent": principle_,
                        "stage": "indicators",
                    }
                    yield outcome_
                    outcome_ = {
                        **outcome,
                        "type": "outcome",
                        "code": outcome_code,
                        "short_name": f"{self.get_framework_id()}_confirmation_{outcome_code}",
                        "parent": principle_,
                        "stage": "confirmation",
                    }
                    yield outcome_

    def get_sections(self) -> list[dict]:
        return list(filter(lambda x: x["type"] == "objective", self.elements))

    def get_section(self, objective_id: str) -> Optional[dict]:
        return next((x for x in self.get_sections() if x["code"] == objective_id), None)


class CAF32Router(CAFLoader):
    """
    Manages routing and view creation for CAF v3.2 assessments.

    The `CAF32Router` class is responsible for configuring routes, generating URLs, and creating
    corresponding view classes for the CAF (Cyber Assessment Framework) v3.2. It supports integration
    with Django's URL patterns and ensures breadcrumbs and context are created for views. This class
    inherits from `CAFLoader`.

    :ivar exit_url: The URL to redirect to after the assessment sequence completes.
    :type exit_url: str
    """

    logger = logging.getLogger("CAF32Router")

    @staticmethod
    def _build_breadcrumbs(element: CAF32Element) -> list[dict[str, str]]:
        breadcrumbs: list = []
        # We can only build the root breadcrumb here as the rest of it is dependent on the current assessment
        breadcrumbs.insert(0, {"url": reverse_lazy("my-account"), "text": "My account"})
        return breadcrumbs

    def __init__(self, exit_url: str = "index") -> None:
        self.exit_url = exit_url
        super().__init__()

    def get_framework_path(self) -> str:
        return os.path.join(settings.BASE_DIR, "..", "frameworks", "cyber-assessment-framework-v3.2.yaml")

    def get_framework_id(self) -> str:
        return "caf32"

    def _get_success_url(self, element: CAF32Element) -> str:
        """
        Determine the success URL for a form.
        If there's a next URL in the sequence, use that, otherwise use the exit URL.
        """
        current_index = self.elements.index(element)
        if current_index + 1 < len(self.elements):
            return self.elements[current_index + 1]["short_name"]
        else:
            return self.exit_url

    def _create_view_and_url(self, element: CAF32Element, form_class=None) -> None:
        """
        Takes an element from the CAF, the url for the next page in the route and a form class
        to create a view class and add a path for the view to Django's urlpatterns.
        """
        url_path = slugify(f"{element['code']}-{element['title']}")
        extra_context = {
            "title": element.get("title"),
            "description": element.get("description"),
            "breadcrumbs": CAF32Router._build_breadcrumbs(element),
        }
        if element["type"] in ["objective", "principle"]:
            template_name = f"caf/{element['type']}.html"
            class_prefix = f"{self.get_framework_id().capitalize()}{element['type'].capitalize()}View"
            element["view_class"] = create_form_view(
                success_url_name=self._get_success_url(element),
                template_name=template_name,
                class_prefix=class_prefix,
                class_id=element["code"],
                extra_context=extra_context | {"objective_data": element},
            )
            url_to_add = path(
                f"{self.get_framework_id()}/{url_path}/",
                element["view_class"].as_view(),
                name=element["short_name"],
            )
            urls.urlpatterns.append(url_to_add)
        else:
            template_name = f"caf/{element['stage']}.html"
            class_prefix = f"{self.get_framework_id().capitalize()}Outcome{element['stage'].capitalize()}View"
            element["view_class"] = create_form_view(
                success_url_name=self._get_success_url(element),
                template_name=template_name,
                form_class=form_class,
                class_prefix=class_prefix,
                stage=element["stage"],
                class_id=element["code"],
                extra_context=extra_context
                | {
                    "objective_name": f"Objective {element['parent']['parent']['code']} - {element['parent']['parent']['title']}",
                    "objective_code": element["parent"]["parent"]["code"],
                    "outcome": element,
                    "objective_data": element["parent"]["parent"],
                },
            )
            url_to_add = path(
                f"{self.get_framework_id()}/{url_path}/{element['stage']}/",
                element["view_class"].as_view(),
                name=element["short_name"],
            )
            urls.urlpatterns.append(url_to_add)
        self.logger.debug(f"Added {url_to_add}")

    def _process_outcome(self, element) -> None:
        if element.get("stage") == "indicators":
            provider: FieldProvider = OutcomeIndicatorsFieldProvider(element)
            indicators_form = create_form(provider)
            self._create_view_and_url(element, form_class=indicators_form)
        elif element.get("stage") == "confirmation":
            provider = OutcomeConfirmationFieldProvider(element)
            outcome_form = create_form(provider)
            self._create_view_and_url(element, form_class=outcome_form)

    def _create_route(self) -> None:
        for element in self.elements:
            if element["type"] == "objective":
                self._create_view_and_url(element, "objective")
            elif element["type"] == "principle":
                self._create_view_and_url(element, "principle")
            elif element["type"] == "outcome":
                self._process_outcome(element)

    # Keeping this interface so we can separate generating the order of the elements
    # from creating the Django urls
    def execute(self) -> None:
        self._create_route()


class CAF40Router(CAFLoader):
    logger = logging.getLogger("CAF40Router")

    def get_framework_path(self) -> str:
        return os.path.join(settings.BASE_DIR, "..", "frameworks", "cyber-assessment-framework-v4.0.yaml")

    def get_framework_id(self) -> str:
        return "caf40"

    def execute(self) -> None:
        """
        Not implemented yet
        :return:
        """
        return None


class CAF32ExcelExporter(CAFLoader):
    """Exports CAF v3.2 framework to a formatted Excel workbook.

    The exporter creates one worksheet per Objective and renders all Principles and
    Outcomes beneath, including indicator rows and data validation lists for answers.
    """

    logger = logging.getLogger("CAF32ExcelExporter")

    # ---- FrameworkLoader hooks -------------------------------------------------
    def get_framework_path(self) -> str:
        return os.path.join(settings.BASE_DIR, "..", "frameworks", "cyber-assessment-framework-v3.2.yaml")

    def get_framework_id(self) -> str:
        return "caf32"

    # ---- Helpers ---------------------------------------------------------------

    def _write_top_header(self, ws) -> int:
        """Write the required instruction cells at the very top of a worksheet.
        Returns the next row index to continue rendering (1-based).
        """
        # Cache styles used repeatedly
        fills = self._fills()
        border = self._thin_border()

        # Columns C..I are used everywhere else; keep the same for header
        for col, width in (("C", 60), ("D", 10), ("E", 60), ("F", 10), ("G", 60), ("H", 10), ("I", 60)):
            ws.column_dimensions[col].width = width

        # Content constants
        title = "PLEASE ENTER CLASSIFICATION (OFFICIAL IF BLANK)"
        instructions = (
            "This is not a substitution for using WebCAF. Unless otherwise agreed with GSG, you should be using WebCAF for creating and submitting assessments under GovAssure.\n"
            "However, you can use this spreadsheet to draft your answers. Contributing outcomes, IGPs and supplementary questions are identical to WebCAF.\n\n"
            'To complete this spreadsheet, provide an answer to each Indicator of Good Practice (IGP) by selecting the appropriate value in the dropdowns adjacent to the "Not achieved", "Partially achieved" and "Achieved" columns. Provide a summary of your evidence for each group of IGPs in column I.\n\n'
            "For each Contributing Outcome, select a dropdown for the achievement, and provide comments justifying the achievement selected.\n\n"
            "For certain Contributing Outcomes, there are supplementary questions which are not part of the CAF but provide additional context to your answers. \n\n"
            "Here are links to WebCAF, GovAssure Stage 3 self-assessment guidance and the five lens mapping model."
        )
        links = (
            ("Five Lens Mapping Model", "PROVIDE LINK ONCE NEW FIVE LENS MODEL IS UPLOADED"),
            (
                "Stage 3 Self-Assessment Guidance",
                "https://www.security.gov.uk/policy-and-guidance/govassure/stage-3-self-assessment/",
            ),
            ("WebCAF", "https://webcaf.service.security.gov.uk/"),
        )

        row = 1
        # Title line
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
        cell = ws.cell(row=row, column=3, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fills["blue"]
        cell.border = border
        row += 1

        # Instruction paragraph (multi-line). Merge across C..I and wrap text.
        ws.merge_cells(start_row=row, start_column=3, end_row=row + 5, end_column=9)
        cell = ws.cell(row=row, column=3, value=instructions)
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        cell.border = border
        row += 6

        # Resource Links header
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
        cell = ws.cell(row=row, column=3, value="Resource Links")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fills["blue"]
        cell.border = border
        row += 1

        # Links rows
        for text, target in links:
            left = ws.cell(row=row, column=3, value=text)
            left.border = border
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=9)
            right = ws.cell(row=row, column=5, value=target)
            right.border = border
            row += 1

        # System name prompt
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=3, value="Please enter name of system being assessed:")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fills["blue"]
        cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        cell.border = border
        cell = ws.cell(row=row, column=9, value="")
        cell.border = border
        row += 2

        return row

    @staticmethod
    def _thin_border() -> Border:
        side = Side(border_style="thin", color="000000")
        return Border(left=side, right=side, top=side, bottom=side)

    @staticmethod
    def _fills() -> dict[str, PatternFill]:
        return {
            "yellow": PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid"),
            "blue": PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid"),
            "green": PatternFill(start_color="C6E2B3", end_color="C6E2B3", fill_type="solid"),
            "pink": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),
            "grey": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
        }

    @staticmethod
    def _validators() -> dict[str, DataValidation]:
        # Common choice list used across achievement columns
        formula = '"agreed,not_true_have_justification,not_true_no_justification"'
        not_achieved_formula = '"true_have_justification,agreed,not_true_no_justification"'
        return {
            key: DataValidation(
                type="list",
                formula1=formula if key in ["partially-achieved", "achieved"] else not_achieved_formula,
                allow_blank=False,
            )
            for key in ("not-achieved", "partially-achieved", "achieved")
        }

    @staticmethod
    def _header_specs(fills: dict[str, PatternFill]) -> list[tuple[str, Optional[PatternFill]]]:
        return [
            ("Achieved", fills["green"]),
            ("Answer", fills["green"]),
            ("Partially Achieved", fills["yellow"]),
            ("Answer", fills["yellow"]),
            ("Not Achieved", fills["pink"]),
            ("Answer", fills["pink"]),
            ("Please summarize your evidence", None),
        ]

    @staticmethod
    def _confirmation_status_validatos() -> dict[str, DataValidation]:
        return {
            "with-partial": DataValidation(
                type="list", formula1='"Achieved,Partially achieved,Not achieved"', allow_blank=False
            ),
            "without-partial": DataValidation(type="list", formula1='"Achieved,Not achieved"', allow_blank=False),
        }

    # ---- Public API ------------------------------------------------------------
    def execute(self) -> Workbook:
        """Build and return the Excel workbook for the framework."""
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        border = self._thin_border()
        fills = self._fills()
        validators = self._validators()
        confirmation_validators = self._confirmation_status_validatos()
        headers = self._header_specs(fills)

        # Iterate objectives -> principles -> outcomes
        for obj_code, obj_data in self.framework["objectives"].items():
            ws = wb.create_sheet(title=f"CAF - Objective {obj_code}")

            # Top header block required by specification
            row = self._write_top_header(ws)

            # Register data validations on the worksheet
            for validator in validators.values():
                ws.add_data_validation(validator)
            for validator in confirmation_validators.values():
                ws.add_data_validation(validator)

            # Set column widths (ensure consistent with header)
            for col, width in (("C", 60), ("D", 10), ("E", 60), ("F", 10), ("G", 60), ("H", 10), ("I", 60)):
                ws.column_dimensions[col].width = width
            # Objective heading
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=3, value=f"Objective {obj_data['code']} - {obj_data['title']}")
            cell.font = Font(bold=True, size=16)
            row += 1

            # Objective description
            ws.merge_cells(start_row=row, start_column=3, end_row=row + 1, end_column=8)
            cell = ws.cell(row=row, column=3, value=obj_data["description"])
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            row += 2

            # Principles
            for _, principle_data in obj_data.get("principles", {}).items():
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
                cell = ws.cell(row=row, column=3, value=f"{principle_data['code']} - {principle_data['title']}")
                cell.font = Font(bold=True, size=14)
                row += 1

                ws.merge_cells(start_row=row, start_column=3, end_row=row + 1, end_column=8)
                cell = ws.cell(row=row, column=3, value=principle_data["description"])
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                row += 3

                # Outcomes
                for _, outcome_data in principle_data.get("outcomes", {}).items():
                    # Outcome header bar
                    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
                    cell = ws.cell(row=row, column=3, value=f"{outcome_data['code']} - {outcome_data['title']}")
                    cell.font = Font(bold=True, size=14, color="FFFFFF")
                    cell.fill = fills["blue"]
                    cell.border = border
                    row += 1

                    # Outcome description bar
                    ws.merge_cells(start_row=row, start_column=3, end_row=row + 1, end_column=9)
                    cell = ws.cell(row=row, column=3, value=outcome_data["description"])
                    cell.font = Font(color="FFFFFF")
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.fill = fills["blue"]
                    cell.border = border
                    row += 2

                    # Column headers
                    for col_idx, (title, fill) in enumerate(headers, start=3):
                        cell = ws.cell(row=row, column=col_idx, value=title)
                        cell.font = Font(bold=True, size=12)
                        cell.border = border
                        if fill:
                            cell.fill = fill
                    row += 1

                    # Indicators block
                    indicators = outcome_data.get("indicators", {})
                    max_len = max((len(v) for v in indicators.values() if isinstance(v, dict)), default=0)

                    for idx in range(max_len):
                        col_idx = 3
                        for key in ("achieved", "partially-achieved", "not-achieved"):
                            values = indicators.get(key, {})
                            if idx < len(values):
                                item_code, item_data = list(values.items())[idx]
                                desc = f"{item_code} - {item_data['description']}"
                                cell = ws.cell(row=row, column=col_idx, value=desc)
                                cell.alignment = Alignment(wrap_text=True)
                                cell.border = border
                                # Fill per column type
                                cell.fill = (
                                    fills["pink"]
                                    if key == "not-achieved"
                                    else fills["yellow"]
                                    if key == "partially-achieved"
                                    else fills["green"]
                                )
                            else:
                                cell = ws.cell(row=row, column=col_idx, value="")
                                cell.fill = fills["grey"]
                                cell.border = border
                            col_idx += 1

                            # Adjacent answer dropdown cell
                            ans_cell = ws.cell(row=row, column=col_idx)
                            ans_cell.border = border
                            validators[key].add(ws[ans_cell.coordinate])
                            col_idx += 1

                        # Evidence cell at the end
                        ev_cell = ws.cell(row=row, column=col_idx, value="")
                        ev_cell.border = border
                        row += 1

                    # Contributing outcome achievement fields
                    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
                    cell = ws.cell(row=row, column=7, value="Contributing Outcome achievement fff:")
                    cell.font = Font(bold=True)
                    cell.border = border

                    cell = ws.cell(
                        row=row,
                        column=9,
                    )
                    if indicators.get("partially-achieved"):
                        validator = confirmation_validators["with-partial"]
                    else:
                        validator = confirmation_validators["without-partial"]
                    validator.add(ws[cell.coordinate])
                    cell.border = border
                    row += 1

                    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
                    cell = ws.cell(
                        row=row,
                        column=7,
                        value=("Please provide comments justifying your achievement for this Contributing Outcome:"),
                    )
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                    cell = ws.cell(row=row, column=9, value="")
                    cell.border = border
                    row += 5

        return wb
