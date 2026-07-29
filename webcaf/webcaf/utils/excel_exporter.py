"""Builds the downloadable Excel assessment template for a CAF framework.

The workbook mirrors the structure of the framework YAML: one worksheet per
Objective with all Principles and Outcomes beneath, including indicator rows
and data validation lists for answers. A hidden mapping sheet
(``JSON_MAP_SHEET_NAME``) records which visible cell feeds which JSON path so
the importer can convert a completed workbook back into assessment JSON.
"""

import logging
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from webcaf.webcaf.utils.excel_importer import JSON_MAP_HEADERS, JSON_MAP_SHEET_NAME

logger = logging.getLogger(__name__)

_COLUMN_WIDTHS = (("C", 60), ("D", 10), ("E", 60), ("F", 10), ("G", 60), ("H", 10), ("I", 60))


def create_assessment_template_workbook(framework_id: str) -> Workbook:
    """Build the Excel template for the framework with the given id (e.g. ``"caf32"``).

    The framework definition is taken from the already-initialised router
    registry, so the YAML file is not re-read from disk.
    """
    # Imported here rather than at module level: the router registry is only
    # populated once the Django app registry is ready.
    from webcaf.webcaf.frameworks import routers

    return build_assessment_template_workbook(routers[framework_id].framework)


def build_assessment_template_workbook(framework: dict[str, Any]) -> Workbook:
    """Build and return the Excel template workbook for the given framework data."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet
    json_map_rows: list[list[Any]] = [JSON_MAP_HEADERS]

    border = _thin_border()
    fills = _fills()
    validators = _validators()
    confirmation_validators = _confirmation_status_validators()
    headers = _header_specs(fills)

    # Iterate objectives -> principles -> outcomes
    for obj_code, obj_data in framework["objectives"].items():
        ws = wb.create_sheet(title=f"CAF - Objective {obj_code}")

        # Top header block required by specification
        row = _write_top_header(ws)

        # Register data validations on the worksheet
        for validator in validators.values():
            ws.add_data_validation(validator)
        for validator in confirmation_validators.values():
            ws.add_data_validation(validator)

        # Set column widths (ensure consistent with header)
        for col, width in _COLUMN_WIDTHS:
            ws.column_dimensions[col].width = width
        # Objective heading
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=3, value=f"Objective {obj_data['code']} - {obj_data['title']}")
        cell.font = Font(bold=True, size=16)
        row += 1

        # Objective description
        ws.merge_cells(start_row=row, start_column=3, end_row=row + 1, end_column=8)
        cell = ws.cell(row=row, column=3, value=obj_data["description"])
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 2

        # Principles
        for _, principle_data in obj_data.get("principles", {}).items():
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=3, value=f"{principle_data['code']} - {principle_data['title']}")
            cell.font = Font(bold=True, size=14)
            row += 1

            ws.merge_cells(start_row=row, start_column=3, end_row=row + 1, end_column=8)
            cell = ws.cell(row=row, column=3, value=principle_data["description"])
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            row += 3

            # Outcomes
            for _, outcome_data in principle_data.get("outcomes", {}).items():
                # Outcome header bar
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
                cell = ws.cell(row=row, column=3, value=f"{outcome_data['code']} - {outcome_data['title']}")
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.fill = fills["blue"]
                cell.border = border
                row += 1

                # Outcome description bar
                ws.merge_cells(start_row=row, start_column=3, end_row=row + 1, end_column=9)
                cell = ws.cell(row=row, column=3, value=outcome_data["description"])
                cell.font = Font(color="FFFFFF")
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.fill = fills["blue"]
                cell.border = border
                row += 2

                # Column headers
                for col_idx, (title, fill) in enumerate(headers, start=3):
                    cell = ws.cell(row=row, column=col_idx, value=title)
                    cell.font = Font(bold=True, size=12)
                    cell.border = border
                    if fill:
                        cell.fill = fill
                row += 1

                # Indicators block
                indicators = outcome_data.get("indicators", {})
                row = _write_indicator_rows(
                    ws, indicators, outcome_data["code"], row, border, fills, validators, json_map_rows
                )

                # Contributing outcome achievement fields
                ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
                cell = ws.cell(row=row, column=7, value="Contributing Outcome achievement:")
                cell.font = Font(bold=True)
                cell.border = border

                cell = ws.cell(
                    row=row,
                    column=9,
                )
                if indicators.get("partially-achieved"):
                    validator = confirmation_validators["with-partial"]
                else:
                    validator = confirmation_validators["without-partial"]
                validator.add(ws[cell.coordinate])
                cell.border = border
                _append_json_map(
                    json_map_rows,
                    ws,
                    cell.coordinate,
                    f"/{outcome_data['code']}/confirmation/outcome_status",
                    "outcome_status",
                    required=True,
                )
                row += 1

                ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
                cell = ws.cell(
                    row=row,
                    column=7,
                    value=("Please provide comments justifying your achievement for this Contributing Outcome:"),
                )
                cell.font = Font(bold=True)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                cell = ws.cell(row=row, column=9, value="")
                cell.border = border
                _append_json_map(
                    json_map_rows,
                    ws,
                    cell.coordinate,
                    f"/{outcome_data['code']}/confirmation/confirm_outcome_confirm_comment",
                    "text",
                )
                row += 5

    map_ws = wb.create_sheet(title=JSON_MAP_SHEET_NAME)
    for map_row in json_map_rows:
        map_ws.append(map_row)
    map_ws.sheet_state = "veryHidden"

    return wb


def _write_top_header(ws) -> int:
    """Write the required instruction cells at the very top of a worksheet.
    Returns the next row index to continue rendering (1-based).
    """
    # Cache styles used repeatedly
    fills = _fills()
    border = _thin_border()

    # Columns C..I are used everywhere else; keep the same for header
    for col, width in _COLUMN_WIDTHS:
        ws.column_dimensions[col].width = width

    # Content constants
    title = "PLEASE ENTER CLASSIFICATION (OFFICIAL IF BLANK)"
    instructions = (
        "This is not a substitution for using WebCAF. Unless otherwise agreed with GSG, you should be using WebCAF for creating and submitting assessments under GovAssure.\n"
        "However, you can use this spreadsheet to draft your answers. Contributing outcomes, IGPs and supplementary questions are identical to WebCAF.\n\n"
        'To complete this spreadsheet, provide an answer to each Indicator of Good Practice (IGP) by selecting the appropriate value in the dropdowns adjacent to the "Not achieved", "Partially achieved" and "Achieved" columns. Provide a summary of your evidence for each group of IGPs in column I.\n\n'
        "For each Contributing Outcome, select a dropdown for the achievement, and provide comments justifying the achievement selected.\n\n"
        "For certain Contributing Outcomes, there are supplementary questions which are not part of the CAF but provide additional context to your answers. \n\n"
        "Here are links to WebCAF, GovAssure Stage 3 self-assessment guidance and the five lens mapping model."
    )
    links = (
        ("Five Lens Mapping Model", "PROVIDE LINK ONCE NEW FIVE LENS MODEL IS UPLOADED"),
        (
            "Stage 3 Self-Assessment Guidance",
            "https://www.security.gov.uk/policy-and-guidance/govassure/stage-3-self-assessment/",
        ),
        ("WebCAF", "https://webcaf.service.security.gov.uk/"),
    )

    row = 1
    # Title line
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=3, value=title)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = fills["blue"]
    cell.border = border
    row += 1

    # Instruction paragraph (multi-line). Merge across C..I and wrap text.
    ws.merge_cells(start_row=row, start_column=3, end_row=row + 5, end_column=9)
    cell = ws.cell(row=row, column=3, value=instructions)
    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    cell.border = border
    row += 6

    # Resource Links header
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=3, value="Resource Links")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = fills["blue"]
    cell.border = border
    row += 1

    # Links rows
    for text, target in links:
        left = ws.cell(row=row, column=3, value=text)
        left.border = border
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=9)
        right = ws.cell(row=row, column=5, value=target)
        right.border = border
        row += 1

    # System name prompt
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=3, value="Please enter name of system being assessed:")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = fills["blue"]
    cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    cell.border = border
    cell = ws.cell(row=row, column=9, value="")
    cell.border = border
    row += 2

    return row


def _write_indicator_rows(
    ws,
    indicators: dict[str, Any],
    outcome_code: str,
    row: int,
    border: Border,
    fills: dict[str, PatternFill],
    validators: dict[str, DataValidation],
    json_map_rows: list[list[Any]],
) -> int:
    """Write the indicator statement/answer grid for an outcome; return the next free row."""
    max_len = max((len(v) for v in indicators.values() if isinstance(v, dict)), default=0)

    for idx in range(max_len):
        col_idx = 3
        for key in ("achieved", "partially-achieved", "not-achieved"):
            values = indicators.get(key, {})
            item_code = None
            if idx < len(values):
                item_code, item_data = list(values.items())[idx]
                desc = f"{item_code} - {item_data['description']}"
                cell = ws.cell(row=row, column=col_idx, value=desc)
                cell.alignment = Alignment(wrap_text=True)
                cell.border = border
                # Fill per column type
                cell.fill = (
                    fills["pink"]
                    if key == "not-achieved"
                    else fills["yellow"] if key == "partially-achieved" else fills["green"]
                )
            else:
                cell = ws.cell(row=row, column=col_idx, value="")
                cell.fill = fills["grey"]
                cell.border = border
            col_idx += 1

            # Adjacent answer dropdown cell
            ans_cell = ws.cell(row=row, column=col_idx)
            ans_cell.border = border
            validators[key].add(ws[ans_cell.coordinate])
            if item_code:
                _append_json_map(
                    json_map_rows,
                    ws,
                    ans_cell.coordinate,
                    f"/{outcome_code}/indicators/{key}_{item_code}",
                    "indicator_answer",
                )
            col_idx += 1

        # Evidence cell at the end
        ev_cell = ws.cell(row=row, column=col_idx, value="")
        ev_cell.border = border
        row += 1

    return row


def _thin_border() -> Border:
    side = Side(border_style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _fills() -> dict[str, PatternFill]:
    return {
        "yellow": PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid"),
        "blue": PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid"),
        "green": PatternFill(start_color="C6E2B3", end_color="C6E2B3", fill_type="solid"),
        "pink": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),
        "grey": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
    }


def _validators() -> dict[str, DataValidation]:
    # Common choice list used across achievement columns
    formula = '"agreed,not_true_have_justification,not_true_no_justification"'
    not_achieved_formula = '"true_have_justification,agreed,not_true_no_justification"'
    return {
        key: DataValidation(
            type="list",
            formula1=formula if key in ["partially-achieved", "achieved"] else not_achieved_formula,
            allow_blank=False,
        )
        for key in ("not-achieved", "partially-achieved", "achieved")
    }


def _header_specs(fills: dict[str, PatternFill]) -> list[tuple[str, Optional[PatternFill]]]:
    return [
        ("Achieved", fills["green"]),
        ("Answer", fills["green"]),
        ("Partially Achieved", fills["yellow"]),
        ("Answer", fills["yellow"]),
        ("Not Achieved", fills["pink"]),
        ("Answer", fills["pink"]),
        ("Please summarize your evidence", None),
    ]


def _confirmation_status_validators() -> dict[str, DataValidation]:
    return {
        "with-partial": DataValidation(
            type="list", formula1='"Achieved,Partially achieved,Not achieved"', allow_blank=False
        ),
        "without-partial": DataValidation(type="list", formula1='"Achieved,Not achieved"', allow_blank=False),
    }


def _append_json_map(
    json_map_rows: list[list[Any]],
    ws,
    cell_coordinate: str,
    json_path: str,
    value_type: str,
    required: bool = False,
) -> None:
    json_map_rows.append([ws.title, cell_coordinate, json_path, value_type, required])
