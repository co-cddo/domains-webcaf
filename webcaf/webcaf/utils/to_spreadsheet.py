from io import BytesIO
from typing import Any, Literal, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from webcaf.webcaf.caf.util import IndicatorStatusChecker
from webcaf.webcaf.models import Assessment, Review, Tip
from webcaf.webcaf.utils.review import get_review_recommendations

MIN_WIDTH = 20
PADDING = 2


def _add_actioned_tab(wb: Workbook, tip: Tip, context: dict[str, Any]) -> None:
    ws = wb.create_sheet("Actioned recommendations")
    ws.append(
        [
            "Type",
            "Contributing outcome",
            "Associated risk",
            "Reviewer recommendation",
            "Recommendation and risk reviewed",
            "Status",
            "Action",
            "Action owner",
            "Resource available",
            "Budget available",
            "Estimated completion date",
            "Reason for no completion date",
        ]
    )
    _set_header_properties(ws, [10, 20, 50, 50, 10, 30, 50, 20, 20, 20, 20, 50])
    for recommendation_type in ["priority_recommendations", "other_recommendations"]:
        for recommendation, group, action in context.get(recommendation_type, []):
            if action and action.action_type == "action_planned":
                action_details = action.action_details
                ws.append(
                    [
                        action.recommendation_category.capitalize(),
                        f"{recommendation.outcome} {recommendation.outcome_title}",
                        f"{'RP' if action.recommendation_category == 'priority' else 'RO'}{group.group_index} — {group.title}",
                        recommendation.id + " - " + recommendation.text,
                        action.recommendation_reviewed.capitalize(),
                        "Action planned",
                        action_details.get("action_taken_description", ""),
                        action_details.get("action_owner", ""),
                        action_details.get("resources_available", ""),
                        action_details.get("budget_available", ""),
                        f"{action_details.get('target_day_day', '')}/{action_details.get('target_day_month', '')}/{action_details.get('target_day_year', '')}"
                        if action_details.get("target_date_provided", "no") == "yes"
                        else "No target date",
                        action_details.get("target_date_unavailable_reason", "N/A")
                        if action_details.get("target_date_provided", "yes") == "no"
                        else "N/A",
                    ]
                )
                _wrap_row_text(ws)


def _add_not_actioned_tab(wb: Workbook, tip: Tip, context: dict[str, Any]) -> None:
    ws = wb.create_sheet("Not actioned recommendations")
    ws.append(
        [
            "Type",
            "Contributing outcome",
            "Associated risk",
            "Reviewer recommendation",
            "Recommendation and risk reviewed",
            "Status",
        ]
    )
    _set_header_properties(ws, [10, 20, 50, 50, 10, 30, 50])
    for recommendation_type in ["priority_recommendations", "other_recommendations"]:
        for recommendation, group, action in context.get(recommendation_type, []):
            if action and action.action_type == "action_not_planned":
                ws.append(
                    [
                        action.recommendation_category.capitalize(),
                        f"{recommendation.outcome} {recommendation.outcome_title}",
                        f"{'RP' if action.recommendation_category == 'priority' else 'RO'}{group.group_index} — {group.title}",
                        recommendation.id + " - " + recommendation.text,
                        action.recommendation_reviewed.capitalize(),
                        "No action planned",
                    ]
                )
                _wrap_row_text(ws)


def tip_to_excel(tip: Tip, context: dict[str, Any]) -> bytes | None:
    """
    Converts the provided `Tip` object into an Excel workbook, including metadata,
    actioned, and not-actioned data, then returns the workbook as a byte stream.
    This function is used to export the `Tip` data into a structured file format
    for external consumption.

    :param tip: The `Tip` object containing the data to be exported to the Excel
        workbook.
    :type tip: Tip
    :param context: A dictionary containing additional context or configuration
        parameters required for creating the workbook. Specific keys and values
        depend on the implementation.
    :type context: dict[str, Any]
    :return: The generated Excel workbook as a byte stream, or None if the export
        cannot be completed.
    :rtype: bytes | None
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws = _add_metadata_tab(wb, tip.review.assessment)
    ws.append(["Status", "Submitted" if tip.is_submitted or tip.is_approved else "Draft"])
    _add_actioned_tab(wb, tip, context)
    _add_not_actioned_tab(wb, tip, context)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def review_to_tip_template_excel(review: Review) -> bytes | None:
    """
    Convert a review object into a TIP template Excel file.

    This function processes the provided `Review` object and generates an Excel
    file compliant with the TIP template format. It creates multiple tabs
    within the workbook, filling them with relevant data from the given review.
    The generated Excel file is returned as a byte stream. If the review has
    not been marked as complete, the function will return `None`.

    :param review: Review object containing data to be exported into the
        TIP template. Must have a complete review to generate the file.
    :type review: Review
    :return: Byte stream representation of the generated TIP template Excel file
        or `None` if the review is incomplete.
    :rtype: bytes or None
    """
    if not review.is_review_complete():
        return None

    wb = Workbook()
    wb.remove(wb.active)

    _add_metadata_tab(wb, review.assessment)
    _add_indicator_tab(wb, review)
    _add_outcome_summary_tab(wb, review)
    _add_recommendations_tab(wb, review, "priority")
    _add_recommendations_tab(wb, review, "normal")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def review_to_excel(review: Review) -> bytes | None:
    """
    Converts a review object into an Excel workbook represented as a binary stream
    of bytes. The workbook will only be generated if the review is marked as
    complete. The resulting Excel file includes multiple tabs reflecting different
    aspects of the review, such as metadata, indicators, outcome summary, and
    recommendations.

    :param review: The review instance containing the data to be exported into an
        Excel workbook.
    :type review: Review
    :return: A binary stream of the Excel workbook if the review is complete;
        otherwise, None.
    :rtype: bytes | None
    """
    if not review.is_review_complete():
        return None

    wb = Workbook()
    wb.remove(wb.active)

    _add_metadata_tab(wb, review.assessment)
    _add_indicator_tab(wb, review)
    _add_outcome_summary_tab(wb, review)
    _add_recommendations_tab(wb, review)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _add_metadata_tab(wb: Workbook, assessment: Assessment) -> Worksheet:
    """
    Adds a metadata tab to the given Excel workbook. This function creates a new sheet titled
    "Review details" and appends various metadata about the assessment, such as organization
    details, assessment period, review type, framework, CAF version, and assigned target profile. It
    formats the sheet by setting appropriate column widths.

    :param wb:
        The Excel workbook object where the metadata tab will be added.
    :type wb: Workbook
    :param assessment:
        An object containing assessment information, including organizational data,
        review type, framework, and CAF profile details.
    :type assessment: Assessment
    :return:
        Worksheet
    """
    ws = wb.create_sheet("Review details")

    review_type_label = dict(Assessment.REVIEW_TYPE_CHOICES).get(assessment.review_type, assessment.review_type)
    framework_label = dict(Assessment.FRAMEWORK_CHOICES).get(assessment.framework, assessment.framework)
    profile_label = dict(Assessment.PROFILE_CHOICES).get(assessment.caf_profile, assessment.caf_profile)

    ws.append(["Organisation:", assessment.system.organisation.name])
    ws.append(["System name:", assessment.system.name])
    ws.append(["Review year:", assessment.assessment_period])
    ws.append(["Review type:", review_type_label.title()])
    ws.append(["CAF version:", framework_label])
    ws.append(["Assigned target CAF profile:", profile_label])
    _set_header_properties(ws, [30, 40], fix=False, bold=False)
    return ws


def _add_indicator_tab(wb: Workbook, review: Review):
    """
    Adds a new worksheet named "IGPs" to the provided workbook and populates it
    with indicator data. The data is drawn from the assessment and review objects,
    structuring it to include contributing outcomes, indicator details,
    self-assessment values, and review results.

    This function organizes and writes information about objectives, principles,
    outcomes, and indicators based on their hierarchical structure. It reflects
    the self-assessment values and corresponding reviewer evaluations into a
    tabular format.

    :param wb: An instance of `Workbook` where the "IGPs" worksheet will be added.
    :param review: An instance of `Review` that encapsulates review data used
                   for generating the indicator tab.
    :return: None
    """
    ws = wb.create_sheet("IGPs")
    assessment: Assessment = review.assessment
    ws.append(
        [
            "Contributing outcome",
            "IGP",
            "IGP wording",
            "Self-assessment",
            "Self-assessment comments",
            "Review",
        ]
        + (
            [
                "Review comments",
            ]
            if assessment.review_type != "peer_review"
            else []
        )
    )
    _set_header_properties(ws, [50, 30, 50, 20, 50, 10, 50])

    review_data = review.get_assessor_response()

    for objective in assessment.get_all_caf_objectives():
        objective_code = objective["code"]

        for principle in objective.get("principles", {}).values():
            for outcome in principle.get("outcomes", {}).values():
                outcome_code = outcome["code"]
                outcome_title = outcome.get("title", "")
                contributing_outcome = f"{outcome_code} {outcome_title}" if outcome_title else outcome_code

                assessment_section = assessment.get_section_by_outcome_id(outcome_code)
                if not assessment_section:
                    continue

                assessment_indicators = assessment_section.get("indicators", {})

                review_outcome = review_data.get(objective_code, {}).get(outcome_code, {})
                review_indicators = review_outcome.get("indicators", {})

                for level in ["achieved", "partially-achieved", "not-achieved"]:
                    level_indicators = outcome.get("indicators", {}).get(level, {})
                    statement_number = 1

                    if level == "achieved":
                        level_display = "Achieved"
                    elif level == "partially-achieved":
                        level_display = "Partially achieved"
                    else:
                        level_display = "Not achieved"

                    for indicator_id, indicator_data in level_indicators.items():
                        indicator_label = f"{outcome_code} {level_display} statement {statement_number}"
                        indicator_text = indicator_data.get("description", "")

                        prefixed_id = f"{level}_{indicator_id}"
                        statement_number += 1

                        self_assessment_value = assessment_indicators.get(prefixed_id, False)
                        self_assessment_comment = assessment_indicators.get(f"{prefixed_id}_comment", "")

                        review_value = review_indicators.get(prefixed_id, "")
                        review_comment = review_indicators.get(f"{prefixed_id}_comment", "")

                        self_assessment_display = "Y" if self_assessment_value else "N"
                        review_display = "Y" if review_value == "yes" else "N"

                        ws.append(
                            [
                                contributing_outcome,
                                indicator_label,
                                indicator_text,
                                self_assessment_display,
                                self_assessment_comment,
                                review_display,
                            ]
                            + ([review_comment] if assessment.review_type != "peer_review" else [])
                        )
                        _wrap_row_text(ws)


def _add_outcome_summary_tab(wb: Workbook, review: Review):
    """
    Adds a new sheet titled "Contributing outcomes" to the given workbook and populates
    it with extracted and processed data from the provided review instance. This function
    is used to summarize contributing outcome data based on assessment and review details.

    :param wb: The workbook object where the new sheet will be added.
    :type wb: Workbook
    :param review: The review object containing assessment data, assessor responses, and
        related information required to generate the contributing outcomes data.
    :type review: Review
    :return: None
    """
    ws = wb.create_sheet("Contributing outcomes")

    ws.append(
        [
            "Contributing outcome",
            "Target CAF profile requirement",
            "Self-assessment status",
            "Review status",
            "Target CAF profile",
        ]
    )
    _set_header_properties(ws, [50, 30, 30, 30, 30])

    assessment: Assessment = review.assessment
    review_data = review.get_assessor_response()

    for objective in assessment.get_all_caf_objectives():
        objective_code = objective["code"]

        for principle in objective.get("principles", {}).values():
            for outcome in principle.get("outcomes", {}).values():
                outcome_code = outcome["code"]
                outcome_title = outcome.get("title", "")
                contributing_outcome = f"{outcome_code} {outcome_title}" if outcome_title else outcome_code

                assessment_section = assessment.get_section_by_outcome_id(outcome_code)
                self_assessment_status = ""
                if assessment_section:
                    confirmation = assessment_section.get("confirmation", {})
                    self_assessment_status = confirmation.get("outcome_status", "")

                review_outcome = review_data.get(objective_code, {}).get(outcome_code, {})
                review_decision = review_outcome.get("review_data", {}).get("review_decision", "")

                if review_decision == "achieved":
                    review_status = "Achieved"
                elif review_decision == "partially-achieved":
                    review_status = "Partially achieved"
                elif review_decision == "not-achieved":
                    review_status = "Not achieved"
                else:
                    review_status = ""

                min_profile_requirement = outcome.get("min_profile_requirement", {})
                target_requirement = min_profile_requirement.get(assessment.caf_profile, "")

                principle_id = outcome_code.rsplit(".", 1)[0]
                status_to_check = review_status if review_status else self_assessment_status
                met_status = IndicatorStatusChecker.indicator_min_profile_requirement_met(
                    assessment, principle_id, outcome_code, status_to_check
                )

                ws.append(
                    [
                        contributing_outcome,
                        target_requirement,
                        self_assessment_status,
                        review_status,
                        "Met" if met_status and met_status == "Yes" else met_status,
                    ]
                )
                _wrap_row_text(ws)


def _build_contributing_outcome_titles(review: Review) -> dict[str, str]:
    """
    Builds a mapping of outcome codes to their display titles ("<code> <title>")
    for every contributing outcome in the review's assessment.

    :param review: The review whose assessment outcomes are read.
    :type review: Review
    :return: Mapping of outcome code to its formatted title.
    :rtype: dict[str, str]
    """
    titles: dict[str, str] = {}
    for objective in review.assessment.get_all_caf_objectives():
        for principle in objective.get("principles", {}).values():
            for outcome in principle.get("outcomes", {}).values():
                outcome_code = outcome["code"]
                outcome_title = outcome.get("title", "")
                titles[outcome_code] = f"{outcome_code} {outcome_title}" if outcome_title else outcome_code
    return titles


def _append_recommendation_rows(
    ws: Worksheet,
    review: Review,
    titles: dict[str, str],
    recommendation_type: Literal["priority", "normal", "all"],
    prefix: str,
    profile_met: str,
    include_risk: bool = True,
) -> None:
    """
    Appends one row per recommendation for the given recommendation type, wrapping
    text on each row as it is added.

    :param ws: The worksheet to append rows to.
    :param review: The review providing the recommendation groups.
    :param titles: Mapping of outcome code to formatted contributing-outcome title.
    :param recommendation_type: Which recommendations to fetch ("priority", "normal" or "all").
    :param prefix: Risk-number prefix (e.g. "RP" or "RO").
    :param profile_met: Target CAF profile value to record for each row.
    :param include_risk: Whether to include the risk number and risk title columns.
    :return: None
    """
    for recommendation_group in get_review_recommendations(review, recommendation_type):
        for recommendation in recommendation_group.recommendations:
            row = [titles[recommendation.outcome], profile_met]
            if include_risk:
                row += [f"{prefix}{recommendation_group.group_index}", recommendation.title]
            row += [recommendation.id, recommendation.text]
            ws.append(row)
            _wrap_row_text(ws)


def _add_recommendations_tab(
    wb: Workbook, review: Review, recommendation_type: Literal["priority", "normal", "all"] = "all"
):
    """
    Adds a "Risks and recommendations" tab to the given workbook based on the provided review information.

    A new sheet is created in the workbook, designed to present recommendations associated
    with outcomes, target profiles, risks, and their contributing factors. The method processes
    assessment and review data to populate the sheet with valuable insights.

    :param wb: The workbook to which the "Risks and recommendations" tab will be added.
    :type wb: Workbook
    :param review: An object containing assessment review data, including assessor responses,
        recommendations, and their associated statuses.
    :type review: Review

    :return: None
    """
    titles = _build_contributing_outcome_titles(review)
    is_peer_review = review.assessment.review_type == "peer_review"

    if recommendation_type == "all":
        ws = wb.create_sheet("Recommendations" if is_peer_review else "Risks and recommendations")
        ws.append(
            ["Contributing outcome", "Target CAF profile"]
            + ([] if is_peer_review else ["Risk number", "Risk"])
            + ["Recommendation number", "Recommendation"]
        )
        _set_header_properties(ws, [40, 20, 20, 70, 20, 70])
        for rec_type, prefix, profile_met in [("priority", "RP", "Not met"), ("normal", "RO", "Met")]:
            _append_recommendation_rows(
                ws,
                review,
                titles,
                cast(Literal["priority", "normal"], rec_type),
                prefix,
                profile_met,
                include_risk=not is_peer_review,
            )
    elif recommendation_type == "priority":
        ws = wb.create_sheet("Priority recommendations")
        ws.append(
            [
                "Contributing outcome",
                "Target CAF profile",
                "Risk number",
                "Risk",
                "Recommendation number",
                "Recommendation",
                "Will you add an action? (Yes/no)",
                "[IF NO] Explain why you will not add an action (Free text - 500 word limit)",
                "[IF YES] What action will you take? (Free text - 500 word limit)",
                "Action owner (Individual or team)",
                "Resources available? (Yes/No, I'm not sure)",
                "Budget available? (Yes/No, I'm not sure)",
                "Target completion date (DD-MM-YYYY)",
                "[IF NO DATE ADDED] Explain why you cannot provide target date (Free text - 500 word limit)",
            ]
        )
        _set_header_properties(ws, [40, 20, 20, 70, 20, 70, 20, 70, 70, 50, 20, 20, 20, 70])
        _append_recommendation_rows(ws, review, titles, "priority", "RP", "Not met")
        _set_col_as_date_format(ws, "M", "dd-mm-yyyy")
    elif recommendation_type == "normal":
        ws = wb.create_sheet("Other recommendations")
        ws.append(
            [
                "",
                "",
                "",
                "",
                "",
                "",
                "ADDING ACTIONS FOR OTHER RECOMMENDATIONS IS OPTIONAL" "",
                "",
                "",
            ]
        )
        _set_header_properties(ws, [40, 20, 20, 70, 20, 70, 20, 70, 50, 20, 20, 20])
        ws.merge_cells("G1:J1")
        ws["G1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        ws.append(
            [
                "Contributing outcome",
                "Target CAF profile",
                "Risk number",
                "Risk",
                "Recommendation number",
                "Recommendation",
                "Will you add an action? (Yes/no)",
                "[IF YES] What action will you take? (Free text - 500 word limit)",
                "Action owner (Individual or team)",
                "Resources available? (Yes/No, I'm not sure)",
                "Budget available? (Yes/No, I'm not sure)",
                "Target completion date (DD-MM-YYYY)",
            ]
        )
        _set_header_properties(ws, [40, 20, 20, 70, 20, 70, 20, 70, 50, 20, 20, 20], row_num=2)
        _append_recommendation_rows(ws, review, titles, "normal", "RO", "Met")
        _set_col_as_date_format(ws, "L", "dd-mm-yyyy")


def _set_header_properties(ws, widths, fix=True, bold=True, row_num=1):
    """
    Sets header properties for the given worksheet. This includes setting column
    widths, applying font styles, text wrapping, and freezing panes above a specified row.

    :param ws: The worksheet object where header properties are to be applied.
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param widths: A list of integers representing the column widths to apply to the header cells.
    :type widths: list[int]
    :param fix: Whether to enable text wrapping and freeze panes. Defaults to True.
    :type fix: bool, optional
    :param bold: Whether to apply bold font to the header cells. Defaults to True.
    :type bold: bool, optional
    :return: None
    """
    for cell, width in zip(ws[row_num], widths):  # row 1
        if bold:
            cell.font = Font(bold=True)
        if fix:
            cell.alignment = Alignment(wrap_text=True)
            #     Freeze everything above A2 - it is always A2 if you want to fix the 1st row
            ws.freeze_panes = "A2"
        col_letter = get_column_letter(cell.column)
        ws.column_dimensions[col_letter].width = max(width, MIN_WIDTH) + PADDING


def _wrap_row_text(ws):
    """
    Adjusts the alignment of cells in the last row of a worksheet to enable text wrapping.
    This is supposed to be used after data is populated into the worksheet for each row.

    This function modifies the alignment of each cell in the last row of the provided worksheet,
    allowing the text within those cells to wrap and fit within the cell boundaries. It determines
    the last row dynamically based on the worksheet's current state.

    :param ws: Worksheet object that needs cell alignment adjustment in its last row.
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :return: None
    """
    row_num = ws.max_row
    for col in ws.columns:
        ws.cell(row=row_num, column=col[0].column).alignment = Alignment(wrap_text=True)


def _set_col_as_date_format(ws: Worksheet, col: str, date_format="yyyy-mm-dd"):
    """
    Sets the date format "yyyy-mm-dd" by default for all cells in the specified column
    of the given worksheet.

    :param ws: The worksheet in which the column's cell formats will be modified.
    :type ws: Worksheet
    :param col: The column identifier (e.g., 'A', 'B', 'C') whose cells
                should be formatted as dates.
    :type col: str
    :param date_format: The date format to be applied to the cells in the specified column.
                        Defaults to "yyyy-mm-dd".
    :type date_format: str
    :return: None
    """
    for cell in ws[col]:
        cell.number_format = date_format
