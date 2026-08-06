"""Converts a completed CAF Excel workbook back into assessment JSON.

Workbooks produced by :mod:`webcaf.webcaf.utils.excel_exporter` contain a
hidden mapping sheet (``JSON_MAP_SHEET_NAME``) whose rows record, for every
answer cell in the visible sheets, the JSON pointer path the value belongs to,
the type of value expected and whether the cell is required. The functions in
this module read that mapping, validate and transform the cell values, and
assemble the nested ``assessments_data`` dictionary stored on an Assessment.
"""

from typing import Any

from openpyxl import load_workbook

JSON_MAP_SHEET_NAME = "__webcaf_json_map"
JSON_MAP_HEADERS = ["visible_sheet", "visible_cell", "json_path", "value_type", "required"]


class ExcelImportError(Exception):
    pass


def excel_to_assessment_json(excel_file) -> dict[str, Any]:
    """Convert an uploaded workbook to assessment JSON, strictly.

    Used when importing for real: raises :class:`ExcelImportError` if any cell
    holds an invalid value *or* if any required cell is blank.

    :param excel_file: A file path or file-like object readable by openpyxl.
    :return: The nested assessment data dictionary, keyed by outcome code.
    """
    data, _raw_rows, missing_required = _parse_excel(excel_file)
    if missing_required:
        raise ExcelImportError("; ".join(f"Required cell {cell} is blank" for cell in missing_required))
    return data


def excel_to_assessment_json_with_raw(excel_file) -> tuple[dict[str, Any], list[tuple[str, str, bool, Any]]]:
    """Convert an uploaded workbook to assessment JSON, leniently, for previewing.

    Unlike :func:`excel_to_assessment_json`, blank required cells do not raise:
    they are simply absent from the returned data, so the preview page can
    highlight them while still showing everything that was parsed.

    :param excel_file: A file path or file-like object readable by openpyxl.
    :return: A tuple of (assessment data dictionary, raw rows). Each raw row is
        a ``(json_path, value_type, required, raw_cell_value)`` tuple, one per
        mapped cell, in mapping-sheet order.
    """
    data, raw_rows, _missing_required = _parse_excel(excel_file)
    return data, raw_rows


def _parse_excel(excel_file) -> tuple[dict[str, Any], list[tuple[str, str, bool, Any]], list[str]]:
    """Parse the Excel workbook into assessment data using its hidden mapping sheet.

    Walks every row of the ``JSON_MAP_SHEET_NAME`` sheet, reads the visible
    cell it points at, transforms the value according to the row's
    ``value_type`` (see :func:`_transform_value`) and writes it into the result
    dictionary at the row's JSON pointer path.

    Invalid values are collected and raised together as one
    :class:`ExcelImportError`. Blank required cells are *not* fatal here: they
    are returned in ``missing_required`` so callers can decide whether to raise
    (strict import) or flag them (preview).

    :param excel_file: A file path or file-like object readable by openpyxl.
    :return: A tuple of (assessment data dictionary, raw rows as described in
        :func:`excel_to_assessment_json_with_raw`, list of blank required cells
        as ``"Sheet!CELL"`` references).
    :raises ExcelImportError: If the workbook has no mapping sheet, the mapping
        sheet's header row is invalid, a mapped sheet is missing, or any cell
        value fails transformation.
    """
    wb = load_workbook(excel_file, data_only=True)
    if JSON_MAP_SHEET_NAME not in wb.sheetnames:
        raise ExcelImportError(f"Workbook is missing the {JSON_MAP_SHEET_NAME} sheet")

    map_ws = wb[JSON_MAP_SHEET_NAME]
    headers = [cell.value for cell in map_ws[1]]
    if headers != JSON_MAP_HEADERS:
        raise ExcelImportError(f"{JSON_MAP_SHEET_NAME} has an invalid header row")

    data: dict[str, Any] = {}
    raw_rows: list[tuple[str, str, bool, Any]] = []
    missing_required: list[str] = []
    errors: list[str] = []

    for visible_sheet, visible_cell, json_path, value_type, required in map_ws.iter_rows(min_row=2, values_only=True):
        if not visible_sheet or not visible_cell or not json_path:
            continue
        if visible_sheet not in wb.sheetnames:
            errors.append(f"Mapped sheet {visible_sheet} does not exist")
            continue

        value = wb[visible_sheet][visible_cell].value
        required_bool = bool(required)
        raw_rows.append((str(json_path), str(value_type), required_bool, value))

        # A blank required cell is not fatal here: record it so the preview can flag the row,
        # but skip transforming/storing it (there is nothing to store).
        if required_bool and _is_blank(value):
            missing_required.append(f"{visible_sheet}!{visible_cell}")
            continue

        try:
            transformed_value = _transform_value(value, str(value_type))
        except ExcelImportError as ex:
            errors.append(f"{visible_sheet}!{visible_cell}: {ex}")
            continue

        _set_json_pointer(data, str(json_path), transformed_value)

    if errors:
        raise ExcelImportError("; ".join(errors))

    _add_confirmation_defaults(data)
    return data, raw_rows, missing_required


def _transform_value(value: Any, value_type: str) -> Any:
    """Validate and convert a raw cell value according to its mapped ``value_type``.

    * ``indicator_answer`` — dropdown answers become booleans (blank counts as
      not agreed / ``False``).
    * ``outcome_status`` — must be one of the achievement statuses; returned as-is.
    * ``text`` — free text; blank becomes an empty string.

    :raises ExcelImportError: If the value is not valid for the type, or the
        type itself is unknown.
    """
    if value_type == "indicator_answer":
        if _is_blank(value):
            return False
        if value in {"agreed", "true_have_justification"}:
            return True
        if value in {"not_true_have_justification", "not_true_no_justification"}:
            return False
        raise ExcelImportError(f"invalid indicator answer {value!r}")

    if value_type == "outcome_status":
        if value in {"Achieved", "Partially achieved", "Not achieved"}:
            return value
        raise ExcelImportError(f"invalid outcome status {value!r}")

    if value_type == "text":
        return "" if _is_blank(value) else str(value)

    raise ExcelImportError(f"unknown value type {value_type!r}")


def _set_json_pointer(data: dict[str, Any], path: str, value: Any) -> None:
    """Write ``value`` into ``data`` at the JSON pointer ``path`` (e.g. ``/A1.a/indicators/x``),
    creating intermediate dictionaries as needed."""
    parts = [part for part in path.split("/") if part]
    if not parts:
        raise ExcelImportError("JSON path cannot be empty")

    current = data
    for part in parts[:-1]:
        current = current.setdefault(part, {})
    current[parts[-1]] = value


def _add_confirmation_defaults(data: dict[str, Any]) -> None:
    """Default each outcome's ``confirm_outcome`` to "confirm", as the form journey would,
    since the spreadsheet has no equivalent input."""
    for outcome_data in data.values():
        if isinstance(outcome_data, dict) and "confirmation" in outcome_data:
            outcome_data["confirmation"].setdefault("confirm_outcome", "confirm")


def _is_blank(value: Any) -> bool:
    return value is None or value == ""
